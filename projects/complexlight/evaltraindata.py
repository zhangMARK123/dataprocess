import json
from matplotlib import type1font
import pandas as pd
import numpy as np
import os
import cv2
import openpyxl
import shutil
####传为pandas参数，分别统计各类别的precious和recall

def plot(type,imgname,pred,gt,score,bbox):
    SHAPE_CLASSES = [
        'circle',
        'uparrow',
        'downarrow',
        'leftarrow',
        'rightarrow',
        'returnarrow',
        'bicycle',   
        'others',
    ]

    COLOR_CLASSES = [
        'red',
        'yellow',
        'green',
        'black',
        'others',
        'unknow'
    
    ]

    TOWARD_CLASSES = [
        'front',
        'side',
        'backside',
        'unknow'
    ]
    CHARACTER_CLASSES=[
        'pass',
        'president',
        'number',
        'word',
        'others',
        'unknow'
    ]
    SIMPLE_CLASSES=['simple',
    'complex']
    root="/disk3/zbh/Datasets/2022_Q1_icu30_test_crop/"
    imgpath=os.path.join(root,imgname)
    img=cv2.imread(imgpath)         
    color=(26,26,139)   
    bbox=bbox.split(",")
    x1=int(float(bbox[0][1:-1]))
    y1=int(float(bbox[1]))
    x2=int(float(bbox[0][1:-1])+float(bbox[2]))
    y2=int(float(bbox[1])+float(bbox[3][:-1]))         
    cv2.rectangle(img,(x1,y1),(x2,y2),color,1)
    type1=type.split("_")[0]
    if type1=="color":
        pred=COLOR_CLASSES[int(pred)]
        gt=COLOR_CLASSES[int(gt)]
    elif type1=="shape":
        pred=SHAPE_CLASSES[int(int(pred))]
        gt=SHAPE_CLASSES[int(gt)]
    elif type1=="character":
        pred=CHARACTER_CLASSES[int(int(pred))]
        gt=CHARACTER_CLASSES[int(gt)]
    elif type1=="toward":
        pred=TOWARD_CLASSES[int(int(pred))]
        gt=TOWARD_CLASSES[int(gt)]
    elif type1=="simple":
        pred=SIMPLE_CLASSES[int(int(pred))]
        gt=SIMPLE_CLASSES[int(gt)]

        
    cv2.putText(img, pred, (15, 15), cv2.FONT_HERSHEY_DUPLEX, 0.5, color)
    cv2.putText(img, gt, (15, 30), cv2.FONT_HERSHEY_DUPLEX, 0.5, color)
    cv2.putText(img, str(score), (15, 45), cv2.FONT_HERSHEY_DUPLEX, 0.5, color)
    
    wpath="/disk3/zs1/Datasets/traindata/"+type
    if not os.path.exists(wpath):
        os.makedirs(wpath)         
    cv2.imwrite(os.path.join(wpath,imgname.split("/")[-1]),img)

def evalresult(labelpath):
    SHAPE_CLASSES = [
        'circle',
        'uparrow',
        'downarrow',
        'leftarrow',
        'rightarrow',
        'returnarrow',
        'bicycle',   
        'others',
    ]

    COLOR_CLASSES = [
        'red',
        'yellow',
        'green',
        'black',
        'others',
        'unknow'
    
    ]

    TOWARD_CLASSES = [
        'front',
        'side',
        'backside',
        'unknow'
    ]
    CHARACTER_CLASSES=[
        'pass',
        'president',
        'number',
        'word',
        'others',
        'unknow'
    ]
    SIMPLE_CLASSES=['simple',
    'complex']
    labels=json.loads(open(labelpath,'r').read())
    # labels_add=json.loads(open(labelpath1,'r').read())
    colorpered=[]
    colorget=[]
    colorhead=[]
    shapepred=[]
    shapegt=[]
    shapehead=[]
    characterpred=[]
    charactergt=[]
    characterhead=[]
    towardpred=[]
    towardgt=[]
    towardhead=[]
    simplepred=[]
    simplegt=[]
    simplehead=[]
    name=[]
    colorscore=[]
    shapescore=[]
    towradscore=[]
    characterscore=[]
    colorname=[]
    shapename=[]
    charactername=[]
    towardname=[]
    simplename=[]
    datacard=[]
    badcase_color=[]
    badcase_shape=[]

    workbook=openpyxl.Workbook()

    temp=labels["objects"]
    for result_output in temp:
        datacard_id=result_output["imgname"].split("/")[0]
        if datacard_id not in datacard:
            datacard.append(datacard_id)
        for i in result_output:
            if i!="bbox" and i.split('_')[-1]!="head" and i!="imgname":
                result_output[i]=float(result_output[i])
        # if result_output["lightboxcolor_head"]=="True":
        #     colorpered.append(result_output["label_color_pred"])
        #     colorget.append(result_output["label_color_gt"])
        #     colorhead.append(result_output['lightboxcolor_head'])
        #     colorscore.append(result_output["score_color_pred"])
        #     colorname.append(result_output["imgname"])
        #     if result_output["label_color_gt"]!=result_output["label_color_pred"] and result_output["label_toward_gt"]==0:
        #     # if result_output["score_color_pred"]<0.3:
        #         # badcase_color.append(result_output["imgname"])
        #     # if result_output["label_color_gt"]!=result_output["label_color_pred"]:
        #         if result_output["label_color_gt"]==0:
        #             plot("color_red",result_output["imgname"],result_output["label_color_pred"],result_output["label_color_gt"],result_output["score_color_pred"],result_output["bbox"])
        #         if result_output["label_color_gt"]==1:
        #             plot("color_yellow",result_output["imgname"],result_output["label_color_pred"],result_output["label_color_gt"],result_output["score_color_pred"],result_output["bbox"])
        #         if result_output["label_color_gt"]==2:
        #             plot("color_gereen",result_output["imgname"],result_output["label_color_pred"],result_output["label_color_gt"],result_output["score_color_pred"],result_output["bbox"])
        #         if result_output["label_color_gt"]==4:
        #             plot("color_other",result_output["imgname"],result_output["label_color_pred"],result_output["label_color_gt"],result_output["score_color_pred"],result_output["bbox"])
        #         if result_output["label_color_gt"]==3:
        #             plot("color_black",result_output["imgname"],result_output["label_color_pred"],result_output["label_color_gt"],result_output["score_color_pred"],result_output["bbox"])
        if result_output["lightboxshape_head"]=="True":
            shapepred.append(result_output["label_shape_pred"])
            shapegt.append(result_output["label_shape_gt"] )
            shapehead.append(result_output['lightboxshape_head'])
            shapescore.append(result_output["score_shape_pred"])
            shapename.append(result_output["imgname"])
            # if result_output["label_shape_pred"]==result_output["label_shape_gt"] and result_output["score_shape_pred"]<0.5 and result_output["label_shape_gt"] in [1,2,3,4,5,6] and result_output["label_toward_gt"]==0:
            if result_output["label_shape_pred"]!=result_output["label_shape_gt"] and result_output["label_toward_gt"]==0:
                # badcase_shape.append(result_output["imgname"])
                if result_output["label_shape_gt"]==0:
                    plot("shape_circle",result_output["imgname"],result_output["label_shape_pred"],result_output["label_shape_gt"],result_output["score_shape_pred"],result_output["bbox"])
                if result_output["label_shape_gt"]==1:
                    plot("shape_up",result_output["imgname"],result_output["label_shape_pred"],result_output["label_shape_gt"],result_output["score_shape_pred"],result_output["bbox"])
                if result_output["label_shape_gt"]==3:
                    plot("shape_left",result_output["imgname"],result_output["label_shape_pred"],result_output["label_shape_gt"],result_output["score_shape_pred"],result_output["bbox"])
                if result_output["label_shape_gt"]==4:
                    plot("shape_right",result_output["imgname"],result_output["label_shape_pred"],result_output["label_shape_gt"],result_output["score_shape_pred"],result_output["bbox"])
                if result_output["label_shape_gt"]==5:
                    plot("shape_return",result_output["imgname"],result_output["label_shape_pred"],result_output["label_shape_gt"],result_output["score_shape_pred"],result_output["bbox"])
                if result_output["label_shape_gt"]==6:
                    plot("shape_bicycle",result_output["imgname"],result_output["label_shape_pred"],result_output["label_shape_gt"],result_output["score_shape_pred"],result_output["bbox"])
    # f=open("hardcase_color2.txt","w")
    # for ii in badcase_color:
    #     f.write(ii)
    #     f.write('\n')
    # f=open("shape2.txt","w")
    # for ii in badcase_shape:
    #     f.write(ii)
    #     f.write('\n')
    
    #     if result_output['character_head']=="True":
    #         characterpred.append(result_output["label_character_pred"] )
    #         charactergt.append(result_output["label_character_gt"])
    #         characterhead.append(result_output['character_head'])
    #         characterscore.append(result_output["score_character_pred"])
    #         charactername.append(result_output["imgname"])
    #     if result_output['toward_head']=="True":
    #         towardpred.append(result_output["label_toward_pred"])
    #         towardgt.append(result_output["label_toward_gt"])
    #         towardhead.append( result_output['toward_head'])
    #         towradscore.append(result_output["score_toward_pred"])
    #         towardname.append(result_output["imgname"])
    #     if result_output['simplelight_head']=="True":
    #         simplepred.append(result_output["label_simplelight_pred"])
    #         simplegt.append(result_output["label_simplelight_gt"])
    #         simplehead.append(result_output['simplelight_head'])
    #         simplename.append(result_output["imgname"])
    # for k in datacard:
    #     workbook=openpyxl.Workbook()
    #     count_color=0
    #     for j in colorname:
    #         if j.split("/")[0]==k:
    #             count_color+=1
    #     sheet1=workbook.active
    #     sheet1.title="color"
    #     sheet1.cell(row=1, column=1, value="name")
    #     sheet1.cell(row=1, column=2, value="score")  
    #     sheet1.cell(row=1, column=3, value="gt")
    #     sheet1.cell(row=1, column=4, value="dt")
    #     sheet1.cell(row=1, column=5, value="true/false")
    #     for i in range(0,count_color):

    #         sheet1.cell(row=i+2, column=1, value=colorname[i].split("/")[-1])
    #         sheet1.cell(row=i+2, column=2, value=colorscore[i])
    #         sheet1.cell(row=i+2, column=3, value=colorget[i])
    #         sheet1.cell(row=i+2, column=4, value=colorpered[i])
    #         # sheet1.cell(row=i+2, column=5, value=color_dt[sort_color[i][0]])
    #     count_shape=0
    #     for j in shapename:
    #         if j.split("/")[0]==k:
    #             count_shape+=1
    #     sheet2=workbook.create_sheet(title="shape")
    #     # index=len(shapepred)
    #     sheet2.cell(row=1, column=1, value="name")
    #     sheet2.cell(row=1, column=2, value="score")  
    #     sheet2.cell(row=1, column=3, value="gt")
    #     sheet2.cell(row=1, column=4, value="dt")
    #     sheet2.cell(row=1, column=5, value="true/false")
    #     for i in range(0,count_shape):
    #         sheet2.cell(row=i+2, column=1, value=shapename[i].split("/")[-1])
    #         sheet2.cell(row=i+2, column=2, value=shapescore[i])
    #         sheet2.cell(row=i+2, column=3, value=shapegt[i])
    #         sheet2.cell(row=i+2, column=4, value=shapepred[i])
    #     sheet3=workbook.create_sheet(title="toward")
    #     count_toward=0
    #     for j in towardname:
    #         if j.split("/")[0]==k:
    #             count_toward+=1
        
    #     sheet3.cell(row=1, column=1, value="name")
    #     sheet3.cell(row=1, column=2, value="score")  
    #     sheet3.cell(row=1, column=3, value="gt")
    #     sheet3.cell(row=1, column=4, value="dt")
    #     sheet3.cell(row=1, column=5, value="true/false")
    #     for i in range(0,count_toward):
    #         sheet3.cell(row=i+2, column=1, value=towardname[i].split("/")[-1])
    #         sheet3.cell(row=i+2, column=2, value=towradscore[i])
    #         sheet3.cell(row=i+2, column=3, value=towardgt[i])
    #         sheet3.cell(row=i+2, column=4, value=towardpred[i])
    #     sheet4=workbook.create_sheet(title="character")
    #     count_character=0
    #     for j in charactername:
    #         if j.split("/")[0]==k:
    #             count_character+=1
    #     # index=len(characterpred)
    #     sheet4.cell(row=1, column=1, value="name")
    #     sheet4.cell(row=1, column=2, value="score")  
    #     sheet4.cell(row=1, column=3, value="gt")
    #     sheet4.cell(row=1, column=4, value="dt")
    #     sheet4.cell(row=1, column=5, value="true/false")
    #     for i in range(0,count_character):
    #         sheet4.cell(row=i+2, column=1, value=charactername[i].split("/")[-1])
    #         sheet4.cell(row=i+2, column=2, value=characterscore[i])
    #         sheet4.cell(row=i+2, column=3, value=charactergt[i])
    #         sheet4.cell(row=i+2, column=4, value=characterpred[i])
    #     workbook.save(k+".xlsx")
            
        
        
        
    


labelpath="../../work_dirs/complexlight_img/onlyhardcase/12hardcase_result.json"
# labelpath1="../../work_dirs/complexlight_img/16_result.json"

evalresult(labelpath)